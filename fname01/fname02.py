
import sqlite3
import os
from . path_my_db import itog_join2, itog2_thermoEx, itog3_FiltrData_Thermal
#db_full_name = itog_join2+os.sep+"gui_05.db"
db_full_name = itog_join2+"gui_07.db"
thermoEx = itog2_thermoEx + "Project_te.exe"
FiltrData_Thermal = itog3_FiltrData_Thermal + "FiltrData.exe"
#"W:\(_work)\_my_lay5\FiltrData_Thermal\FiltrData.exe" 


#conn = sqlite3.connect("gui_05.db")
conn = sqlite3.connect(db_full_name)

#print(f"(это fname02.py) {db_full_name =} ")
#input("=== press key ===")

cursor = conn.cursor()

from pathlib import Path as Path2
import sys
from os import path
import openpyxl
import xlrd

# TODO: со строки 174
#def myfname_experim(fname):  # убрать, т.к. ввод из формы и структура уже другая

def myfname(fname):

    if path.isdir(fname):  # !!! то диалоговое окно вообще-то не даёт
        sg.popup("это директория, а надо выбрать файл", fname, font=("Helvetica", 20))

    elif path.isfile(fname):
        full_name = path.basename(fname)
        name = path.splitext(full_name)[0]
        ext = Path2(full_name).suffixes[-1]
        print(f"{ext=}")
        if ext in ['.csv', '.xlsx', '.txt', '.xls']:
            print("OK, у файла нужное  расширение")
        else:
            sg.theme('HotDogStand')
            sg.popup("======== это не то, что надо!\n(не xlsx/xls/txt/csv)\nваш файл: ",
                     fname, font=("Helvetica", 20))
            sg.theme(mytheme)
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(fname)  # fname - полный путь
            shnames = wb.sheetnames
            sheet01 = wb[shnames[0]]  # Сплавы
            A = sheet01.cell(column=1, row=2).value

            gruppa = []
            for z in range(2, 20):
                if not (sheet01.cell(column=2, row=z).value) is None:
                    gruppa.append((A,  # именно 2 (( надо !!!!!!!!!!!
                                   sheet01.cell(column=2, row=z).value,
                                   sheet01.cell(column=3, row=z).value,
                                   sheet01.cell(column=4, row=z).value,
                                   sheet01.cell(column=5, row=z).value,
                                   )
                                  )  # gruppa.append

            cursor.executemany("""INSERT INTO `st_Сплави` ( 
                             `id_хто_внiс_сплав`,       
                             `Метал`,         --`Название металла`,
                             `Марка`,                  
                             `Класифікація`,           
                             `Доповнення`,           --`Дополнение`  
                             `Де_використовують`          -- Применение   
                             ) \
                             VALUES (1,?,?,?,?,?)
                             """,
                               gruppa)

            conn.commit()
            sheet02 = wb[shnames[1]]  # Химический состав
            #  в базу ======================================================
            gruppa02 = []

            for z in range(2, 20):  ## !!!! OK
                if not (sheet02.cell(column=1, row=z).value) is None:
                    gruppa02.append((  # именно 2 (( надо !!!!!!!!!!!
                        sheet02.cell(column=1, row=z).value,
                        sheet02.cell(column=2, row=z).value,
                        sheet02.cell(column=3, row=z).value,
                        sheet02.cell(column=4, row=z).value,
                        sheet02.cell(column=5, row=z).value,
                        sheet02.cell(column=6, row=z).value,
                        sheet02.cell(column=7, row=z).value,
                        sheet02.cell(column=8, row=z).value
                    )
                    )  # gruppa.append

            cursor.executemany("""INSERT INTO `st_Хiмiчний_склад` ( 

                             `id_хто_внiс_хiм_склад`,       
                             `Марка`, 
                             `Вуглець_С`,
                             `Кремній_Si`,
                             `Марганець_Mn`,
                             `Сірка_S`,
                             `Фосфор_P`,	
                             `Хром_Cr`,	
                             `C_plus_Si`	
                             )
                             VALUES (1,?,?,?,?,?,?,?,?)
                             """,
                               gruppa02)

            conn.commit()
            sheet03 = wb[shnames[2]]  # Механические свойства

            #  в базу ======================================================
            gruppa03 = []  # st_Механiчнi_властивостi

            for z in range(2, 20):  ## !!!! OK
                if not (sheet03.cell(column=1, row=z).value) is None:
                    gruppa03.append((  # именно 2 (( надо !!!!!!!!!!!
                        sheet03.cell(column=1, row=z).value,
                        sheet03.cell(column=2, row=z).value,
                        sheet03.cell(column=3, row=z).value,
                        sheet03.cell(column=4, row=z).value,
                        sheet03.cell(column=5, row=z).value,
                        sheet03.cell(column=6, row=z).value,
                        sheet03.cell(column=7, row=z).value,
                        sheet03.cell(column=8, row=z).value,
                        sheet03.cell(column=9, row=z).value,
                    )
                    )  # gruppa.append

            cursor.executemany("""INSERT INTO `st_Механiчнi_властивостi` ( 

            	`id_хто_внiс_механ_власт`,
                    `Назва`,
                    `Сортамент`,
            	`Межа_короткочасної_міцності_МПа`,
            	`Межа_пропорційності_МПа`,
            	`Відносне_подовження_при_разриві`,
            	`Відносне_звуження`,
            	`Ударна_в_язкість_кДж_м2`, 
            	`Термообробка`, 
                    `Твердість_за_Бринелем_МПа_КЧ30_6_ГОСТ_1215_79`)

                     VALUES (1,?,?,?,?,?,?,?,?,?)
                             """, gruppa03)

            conn.commit()
            sheet04 = wb[shnames[3]]  # Механические свойства
            gruppa04 = []  # st_Механiчнi_властивостi

            for z in range(2, 20):  ## !!!! OK
                if not (sheet04.cell(column=1, row=z).value) is None:
                    gruppa04.append((  # именно 2 (( надо !!!!!!!!!!!
                        sheet04.cell(column=1, row=z).value,
                        sheet04.cell(column=2, row=z).value,
                        sheet04.cell(column=3, row=z).value,
                        sheet04.cell(column=4, row=z).value,
                        sheet04.cell(column=5, row=z).value,
                        sheet04.cell(column=6, row=z).value,
                        sheet04.cell(column=7, row=z).value,
                    )
                    )  # gruppa.append

            cursor.executemany("""INSERT INTO `st_Фiзичнi_властивостi` ( 

            	`id_хто_внiс_фiз_властив`,
            	`Марка`, 
            	---`Температура_при_которой_получены_данные_свойства`,
            	`Модуль_упругості_першого_роду`,
            	`Коефіцієнт_температурного_лінійного_розширення`, 
            	`Коэффициент_теплопроводности_теплоемкость_материала`, 
            	`Щільність матеріалу`,
            	`C_Питома_теплоємність_матеріалу`,
            	`R_10_9_Питомий_електроопір_Om_m`
            	 )
                     VALUES (1,?,?,?,?,?,?,?)
                             """, gruppa04)

            conn.commit()
# end of def myfname(fname):
#  ===============================================================

def myfname_experim(fname):  # убрать, т.к. ввод из формы и структура уже другая


    if path.isdir(fname):  # !!! то диалоговое окно вообще-то не даёт
        sg.popup("это директория, а надо выбрать файл", fname, font=("Helvetica", 20))

    elif path.isfile(fname):
        full_name = path.basename(fname)
        name = path.splitext(full_name)[0]
        ext = Path2(full_name).suffixes[-1]

        print(f"{ext=}")
        if ext in ['.csv', '.xlsx', '.txt', '.xls']:
            print("OK, у файла нужное  расширение")
        else:
            sg.theme('HotDogStand')
            sg.popup("======== это не то, что надо!\n(не xlsx/xls/txt/csv)\nваш файл: ",
                     fname, font=("Helvetica", 20))
            sg.theme(mytheme)

        if ext == '.xlsx':
            wb = openpyxl.load_workbook(fname)  # fname - полный путь
            #sheet01 = wb[name]  # если имя листа совпадает с именем файла
            shnames = wb.sheetnames

            # считал удачно

            sheet01 = wb[shnames[0]]  # Сплавы

            # print( sheet01.cell(column=1, row=2).value )
            A = sheet01.cell(column=1, row=2).value

            #  в базу ======================================================
            gruppa = []
            for z in range(2, 20):
                if not (sheet01.cell(column=2, row=z).value) is None:
                    gruppa.append((A,  # именно 2 (( надо !!!!!!!!!!!
                                   sheet01.cell(column=2, row=z).value,
                                   sheet01.cell(column=3, row=z).value,
                                   sheet01.cell(column=4, row=z).value,
                                   sheet01.cell(column=5, row=z).value,
                                   )
                                  )  # gruppa.append
            cursor.executemany("""INSERT INTO `experim_Сплави` ( 
                             `id_хто_внiс_сплав`,       
                             `Метал`,         --`Название металла`,
                             `Марка`,                  
                             `Класифікація`,           
                             `Доповнення`,           --`Дополнение`  
                             `Де_використовують`          -- Применение   
                             ) \
                             VALUES (1,?,?,?,?,?)
                             """,
                               gruppa)

            conn.commit()

            # ============================================

            sheet02 = wb[shnames[1]]  # Химический состав
            #  в базу ======================================================
            gruppa02 = []

            for z in range(2, 20):  ## !!!! OK
                if not (sheet02.cell(column=1, row=z).value) is None:
                    gruppa02.append((  # именно 2 (( надо !!!!!!!!!!!
                        sheet02.cell(column=1, row=z).value,
                        sheet02.cell(column=2, row=z).value,
                        sheet02.cell(column=3, row=z).value,
                        sheet02.cell(column=4, row=z).value,
                        sheet02.cell(column=5, row=z).value,
                        sheet02.cell(column=6, row=z).value,
                        sheet02.cell(column=7, row=z).value,
                        sheet02.cell(column=8, row=z).value
                    )
                    )  # gruppa.append

            cursor.executemany("""INSERT INTO `experim_Хiмiчний_склад` ( 

                             `id_хто_внiс_хiм_склад`,       
                             `Марка`, 
                             `Вуглець_С`,
                             `Кремній_Si`,
                             `Марганець_Mn`,
                             `Сірка_S`,
                             `Фосфор_P`,	
                             `Хром_Cr`,	
                             `C_plus_Si`	
                             )
                             VALUES (1,?,?,?,?,?,?,?,?)
                             """,
                               gruppa02)

            conn.commit()

            # ==============================================

            # print("====================")
            sheet03 = wb[shnames[2]]  # Механические свойства

            #  в базу ======================================================
            gruppa03 = []  # st_Механiчнi_властивостi

            for z in range(2, 20):  ## !!!! OK
                if not (sheet03.cell(column=1, row=z).value) is None:
                    gruppa03.append((  # именно 2 (( надо !!!!!!!!!!!
                        sheet03.cell(column=1, row=z).value,
                        sheet03.cell(column=2, row=z).value,
                        sheet03.cell(column=3, row=z).value,
                        sheet03.cell(column=4, row=z).value,
                        sheet03.cell(column=5, row=z).value,
                        sheet03.cell(column=6, row=z).value,
                        sheet03.cell(column=7, row=z).value,
                        sheet03.cell(column=8, row=z).value,
                        sheet03.cell(column=9, row=z).value,
                    )
                    )  # gruppa.append

            cursor.executemany("""INSERT INTO `experim_Механiчнi_властивостi` ( 

            	`id_хто_внiс_механ_власт`,
                    `Назва`,
                    `Сортамент`,
            	`Межа_короткочасної_міцності_МПа`,
            	`Межа_пропорційності_МПа`,
            	`Відносне_подовження_при_разриві`,
            	`Відносне_звуження`,
            	`Ударна_в_язкість_кДж_м2`, 
            	`Термообробка`, 
                    `Твердість_за_Бринелем_МПа_КЧ30_6_ГОСТ_1215_79`)

                     VALUES (1,?,?,?,?,?,?,?,?,?)
                             """, gruppa03)

            conn.commit()

            # ==============================================

            # print("====================")

            sheet04 = wb[shnames[3]]  # Механические свойства
            #  в базу ======================================================
            gruppa04 = []  # st_Механiчнi_властивостi

            for z in range(2, 20):  ## !!!! OK
                if not (sheet04.cell(column=1, row=z).value) is None:
                    gruppa04.append((  # именно 2 (( надо !!!!!!!!!!!
                        sheet04.cell(column=1, row=z).value,
                        sheet04.cell(column=2, row=z).value,
                        sheet04.cell(column=3, row=z).value,
                        sheet04.cell(column=4, row=z).value,
                        sheet04.cell(column=5, row=z).value,
                        sheet04.cell(column=6, row=z).value,
                        sheet04.cell(column=7, row=z).value,
                    )
                    )  # gruppa.append

            cursor.executemany("""INSERT INTO `experim_Фiзичнi_властивостi` ( 

            	`id_хто_внiс_фiз_властив`,
            	`Марка`, 
            	---`Температура_при_которой_получены_данные_свойства`,
            	`Модуль_упругості_першого_роду`,
            	`Коефіцієнт_температурного_лінійного_розширення`, 
            	`Коэффициент_теплопроводности_теплоемкость_материала`, 
            	`Щільність матеріалу`,
            	`C_Питома_теплоємність_матеріалу`,
            	`R_10_9_Питомий_електроопір_Om_m`
            	 )
                     VALUES (1,?,?,?,?,?,?,?)
                             """, gruppa04)

            conn.commit()

# end of def myfname_experim




